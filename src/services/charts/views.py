from datetime import datetime

import openpyxl
from django.db.models import Sum, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.views import View
from openpyxl import Workbook
from openpyxl.cell import MergedCell
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from src.core.mixins import AdminRequiredMixin
from src.services.assets.models import AccountBalance
from src.services.expense.models import JournalExpense
from src.services.invoice.models import Invoice
from src.services.project.models import Project
from src.services.transaction.models import Ledger
from src.web.dashboard.utils import ledger_filter
from src.services.expense.forms import DateRangeForm, YearForm


class ChartsIndex(AdminRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        projects = Project.objects.all()
        form = DateRangeForm()
        year_form = YearForm()
        return render(request, "charts/charts_index.html", {"projects": projects, 'form':form, 'years':year_form})

    def post(self, request, *args, **kwargs):
        form = DateRangeForm(request.POST)

        if form.is_valid():
            start_date = form.cleaned_data.get("start_date")
            end_date = form.cleaned_data.get("end_date")

            if start_date and end_date:
                return generate_expense_report(request, start_date, end_date)
            else:
                print("Form is not valid")
            return self.get(request, *args, **kwargs)

def generate_project_report(request, pk):
    # Retrieve the project by primary key
    project = Project.objects.get(pk=pk)
    invoices = project.invoices.all()
    expenses = project.expenses.all()

    # Create a new workbook and set the title
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{project.project_name} Records"

    # Styles
    title_font = Font(size=14, bold=True, color="FFFFFF")
    header_font = Font(size=12, bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    centered_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    title_fill = PatternFill("solid", fgColor="4F81BD")  # Blue background
    header_fill = PatternFill("solid", fgColor="A9C6E8")  # Light blue background
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Helper function to style a cell
    def apply_cell_style(cell, font=None, alignment=None, fill=None, border=None):
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if fill:
            cell.fill = fill
        if border:
            cell.border = border

    # Add Project Title
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    sheet["A1"] = f"{project.project_name} Financial Overview"
    apply_cell_style(
        sheet["A1"], font=title_font, alignment=centered_alignment, fill=title_fill
    )

    # Add Project Information
    current_row = 3
    sheet.append(
        ["Project Name", "Description", "Customer", "Status", "Current Budget", "-"]
    )
    sheet.append(
        [
            project.project_name,
            project.description or "N/A",
            str(project.customer),
            project.project_status,
            project.get_total_budget() or "N/A",
        ]
    )

    # Style the project details
    for row in sheet.iter_rows(
        min_row=current_row, max_row=current_row + 1, min_col=1, max_col=5
    ):
        for cell in row:
            apply_cell_style(
                cell, font=normal_font, alignment=left_alignment, border=thin_border
            )

    # Add "Total Budget Assigned" table
    current_row += 3  # Move down after project details
    sheet.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=2
    )
    sheet[f"A{current_row}"] = "Total Budget Assigned"
    apply_cell_style(
        sheet[f"A{current_row}"],
        font=header_font,
        alignment=centered_alignment,
        fill=title_fill,
    )

    # Table headers
    current_row += 1
    headers = ["Source", "Amount"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(
            cell,
            font=header_font,
            alignment=centered_alignment,
            fill=header_fill,
            border=thin_border,
        )

    # Budget inflow data
    budget_assigned_total = (
        Ledger.objects.filter(
            project_id=pk, transaction_type="BUDGET_ASSIGN"
        ).aggregate(Sum("amount"))["amount__sum"]
        or 0
    )
    loan_created_total = (
        Ledger.objects.filter(project_id=pk, transaction_type="CREATE_LOAN").aggregate(
            Sum("amount")
        )["amount__sum"]
        or 0
    )
    total_inflow = budget_assigned_total + loan_created_total

    data_rows = [
        ("Assigned from Wallet", budget_assigned_total),
        ("Loan", loan_created_total),
        ("Total", total_inflow),
    ]

    # Add data rows
    for row_data in data_rows:
        current_row += 1
        source_cell = sheet.cell(row=current_row, column=1)
        amount_cell = sheet.cell(row=current_row, column=2)
        source_cell.value, amount_cell.value = row_data
        apply_cell_style(
            source_cell, font=normal_font, alignment=left_alignment, border=thin_border
        )
        apply_cell_style(
            amount_cell,
            font=normal_font,
            alignment=centered_alignment,
            border=thin_border,
        )

    # Add Invoices Section
    current_row += 2  # Add spacing
    sheet.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=5
    )
    sheet[f"A{current_row}"] = "Invoices"
    apply_cell_style(
        sheet[f"A{current_row}"],
        font=header_font,
        alignment=centered_alignment,
        fill=title_fill,
    )

    # Table headers for invoices
    current_row += 1
    invoice_headers = [
        "Invoice Number",
        "Client Name",
        "Total Amount",
        "Status",
        "Date",
    ]
    for col, header in enumerate(invoice_headers, start=1):
        cell = sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(
            cell,
            font=header_font,
            alignment=centered_alignment,
            fill=header_fill,
            border=thin_border,
        )

    # Data rows for invoices
    paid_invoices_total = 0
    unpaid_invoices_total = 0

    for invoice in invoices:
        current_row += 1
        sheet.append(
            [
                invoice.invoice_number,
                invoice.client_name,
                invoice.total_amount,
                invoice.status,
                invoice.date.strftime("%Y-%m-%d") if invoice.date else "",
            ]
        )
        # Calculate total amounts based on the status
        if invoice.status.lower() == "paid":
            paid_invoices_total += invoice.total_amount
        else:
            unpaid_invoices_total += invoice.total_amount

    # Add totals for paid and unpaid invoices with better alignment
    current_row += 1  # Move to the next row for totals
    # Total Paid Invoices
    sheet.append(["", "Total Paid Invoices", paid_invoices_total, "", ""])
    apply_cell_style(
        sheet.cell(row=current_row, column=1),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=2),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=3),
        font=normal_font,
        alignment=centered_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=4),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=5),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )

    # Total Unpaid Invoices
    current_row += 1  # Move to the next row for totals
    sheet.append(["", "Total Unpaid Invoices", unpaid_invoices_total, "", ""])
    apply_cell_style(
        sheet.cell(row=current_row, column=1),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=2),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=3),
        font=normal_font,
        alignment=centered_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=4),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=5),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )

    # Add Expenses Section
    current_row += 2  # Add spacing
    sheet.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=6
    )
    sheet[f"A{current_row}"] = "Expenses"
    apply_cell_style(
        sheet[f"A{current_row}"],
        font=header_font,
        alignment=centered_alignment,
        fill=title_fill,
    )

    # Table headers for expenses
    current_row += 1
    expense_headers = [
        "Description",
        "Amount",
        "Source",
        "Category",
        "Payment Status",
        "Vendor",
    ]
    for col, header in enumerate(expense_headers, start=1):
        cell = sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(
            cell,
            font=header_font,
            alignment=centered_alignment,
            fill=header_fill,
            border=thin_border,
        )

    # Data rows for expenses
    paid_expenses_total = 0
    unpaid_expenses_total = 0

    for expense in expenses:
        current_row += 1
        sheet.append(
            [
                expense.description,
                expense.amount,
                expense.budget_source,
                expense.category.name if expense.category else "N/A",
                expense.payment_status,
                expense.vendor.name if expense.vendor else "N/A",
            ]
        )
        # Calculate total amounts based on the payment status
        if expense.payment_status.lower() == "paid":
            paid_expenses_total += expense.amount
        else:
            unpaid_expenses_total += expense.amount

    # Add totals for paid and unpaid expenses with better alignment
    current_row += 1  # Move to the next row for totals
    # Total Paid Expenses
    sheet.append(["", "Total Paid Expenses", paid_expenses_total, "", "", ""])
    apply_cell_style(
        sheet.cell(row=current_row, column=1),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=2),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=3),
        font=normal_font,
        alignment=centered_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=4),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=5),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=6),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )

    # Total Unpaid Expenses
    current_row += 1  # Move to the next row for totals
    sheet.append(["", "Total Unpaid Expenses", unpaid_expenses_total, "", "", ""])
    apply_cell_style(
        sheet.cell(row=current_row, column=1),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=2),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=3),
        font=normal_font,
        alignment=centered_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=4),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=5),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=6),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )

    # Add Trial Balance Section
    current_row += 2  # Add spacing before the new section
    sheet.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=5
    )
    sheet[f"A{current_row}"] = "Trial Balance"
    apply_cell_style(
        sheet[f"A{current_row}"],
        font=header_font,
        alignment=centered_alignment,
        fill=title_fill,
    )

    # Table headers for trial balance
    current_row += 1
    trial_balance_headers = [
        "Budget Assigned (Loan + Wallet)",
        "Expenditure (Loans + Expenses)",
        "Client Funds (Paid)",
        "Client Funds (Unpaid)",
        "Net Total",
    ]
    for col, header in enumerate(trial_balance_headers, start=1):
        cell = sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(
            cell,
            font=header_font,
            alignment=centered_alignment,
            fill=header_fill,
            border=thin_border,
        )

    # Initialize totals
    budget_assigned_total = 0
    expenditure_total = 0
    client_funds_paid = Invoice.calculate_total_received(project_id=project.pk)
    # TODO: FUCKTHISSHIT CODE I forgot to create ledger instances for unpaid invoices so i'll need to query it.
    client_funds_unpaid = Invoice.calculate_total_receivables(project_id=project.pk)

    # Filter ledger entries by project
    ledger_entries = Ledger.objects.filter(project=project)

    # Categorize transactions
    for entry in ledger_entries:
        if (
            entry.transaction_type == "BUDGET_ASSIGN"
            or entry.transaction_type == "CREATE_LOAN"
        ):
            budget_assigned_total += entry.amount
        elif entry.transaction_type == "CREATE_EXPENSE":
            expenditure_total += entry.amount

    # Add row for trial balance totals
    current_row += 1
    # Budget Assigned column
    sheet.cell(row=current_row, column=1, value=budget_assigned_total)
    apply_cell_style(
        sheet.cell(row=current_row, column=1),
        font=normal_font,
        alignment=right_alignment,
        border=thin_border,
    )

    # Expenditure column
    sheet.cell(row=current_row, column=2, value=expenditure_total)
    apply_cell_style(
        sheet.cell(row=current_row, column=2),
        font=normal_font,
        alignment=right_alignment,
        border=thin_border,
    )

    # Client Funds (Paid) column
    sheet.cell(row=current_row, column=3, value=client_funds_paid)
    apply_cell_style(
        sheet.cell(row=current_row, column=3),
        font=normal_font,
        alignment=right_alignment,
        border=thin_border,
    )

    # Client Funds (Unpaid) column
    sheet.cell(row=current_row, column=4, value=f"({client_funds_unpaid})")
    apply_cell_style(
        sheet.cell(row=current_row, column=4),
        font=normal_font,
        alignment=right_alignment,
        border=thin_border,
    )

    # Net Total column (Client Funds - Expenditure)
    net_total = client_funds_paid - expenditure_total
    sheet.cell(row=current_row, column=5, value=net_total)
    apply_cell_style(
        sheet.cell(row=current_row, column=5),
        font=normal_font,
        alignment=right_alignment,
        border=thin_border,
    )

    # Add today's date and time after the trial balance
    current_row += 7  # Add some spacing
    today_date_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.cell(
        row=current_row, column=1, value=f"Report Generated at: {today_date_time}"
    )
    apply_cell_style(
        sheet.cell(row=current_row, column=1),
        font=normal_font,
        alignment=left_alignment,
        border=thin_border,
    )

    # Add the computer-generated report warning
    current_row += 1
    warning_message = "This is a computer-generated report and might be prone to errors or inaccuracies. Please verify the data carefully."
    sheet.cell(row=current_row, column=1, value=warning_message)
    apply_cell_style(
        sheet.cell(row=current_row, column=1),
        font=Font(color="FF0000"),
        alignment=left_alignment,
        border=thin_border,
    )

    # Adjust column widths to fit content
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(
            col[0].column
        )  # Get column letter from the first cell
        for cell in col:
            if cell.value and not isinstance(cell, MergedCell):  # Skip merged cells
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max(max_length + 2, 15)

    # Save workbook to response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{project.project_name}_records.xlsx"'
    )
    workbook.save(response)

    return response

def project_expenses(request, pk):
    # Retrieve the project by primary key
    project = Project.objects.get(pk=pk)

    # Fetch all the ledger entries related to this project
    ledger_entries = Ledger.objects.filter(project=project)

    # Create a new workbook and set the title
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"{project.project_name} Ledger Journal"

    # Styles
    header_font = Font(size=12, bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    centered_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # Blue background for headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Add a title row with project information
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    sheet["A1"] = f"Ledger Journal for {project.project_name}"
    sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    sheet["A1"].alignment = centered_alignment
    sheet["A1"].fill = header_fill

    # Add Column Headers
    sheet.append(
        ["Transaction Type", "Amount", "Source", "Destination", "Reason", "Date"]
    )

    # Apply header styles
    for cell in sheet[2]:
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.fill = header_fill
        cell.border = thin_border

    # Add Ledger Entries
    for ledger in ledger_entries:
        sheet.append(
            [
                ledger.get_transaction_type_display(),  # Human-readable transaction type
                ledger.amount,
                ledger.source,
                ledger.destination,
                ledger.reason if ledger.reason else "N/A",  # Handle missing reason
                ledger.created_at.strftime("%Y-%m-%d %H:%M:%S"),  # Format the date
            ]
        )

    # Apply normal cell styles for the data rows
    for row in sheet.iter_rows(
        min_row=3, max_row=3 + len(ledger_entries), min_col=1, max_col=6
    ):
        for cell in row:
            cell.font = normal_font
            cell.alignment = (
                left_alignment if cell.column != 2 else centered_alignment
            )  # Amount is centered
            cell.border = thin_border

    # Adjust column widths
    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 15
    sheet.column_dimensions["C"].width = 25
    sheet.column_dimensions["D"].width = 25
    sheet.column_dimensions["E"].width = 30
    sheet.column_dimensions["F"].width = 20

    # Return the Excel file as a response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{project.project_name}_ledger_journal.xlsx"'
    )

    # Save the workbook to the response object
    workbook.save(response)

    return response

def generate_bank_statements_view(request):
    # Create a new workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Bank Statements"

    # Define styles
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    accounts = AccountBalance.objects.all()
    start_col = 1  # Starting column for the first account

    for account in accounts:
        # Adjust column widths dynamically for this account's section
        for col_offset in range(7):  # Updated range for extra column
            col_letter = openpyxl.utils.get_column_letter(start_col + col_offset)
            sheet.column_dimensions[col_letter].width = (
                20  # Consistent width for all columns
            )

        # Add Account Header
        sheet.merge_cells(
            start_row=1, start_column=start_col, end_row=1, end_column=start_col + 6
        )  # Updated to include new column
        sheet.cell(row=1, column=start_col).value = account.account_name
        sheet.cell(row=1, column=start_col).font = header_font
        sheet.cell(row=1, column=start_col).alignment = center_align

        # Add Table Headers
        headers = [
            "Sr.No",
            "Date",
            "Description",
            "Transaction Type",  # New header
            "Debit",
            "Credit",
            "Amount",
        ]
        for col_offset, header in enumerate(headers):
            cell = sheet.cell(row=2, column=start_col + col_offset)
            cell.value = header
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style

        # Add Opening Balance
        current_balance = account.starting_balance or 0
        sheet.cell(row=3, column=start_col + 1).value = "1-Jul-22"
        sheet.cell(row=3, column=start_col + 2).value = "Opening Balance"
        sheet.cell(row=3, column=start_col + 6).value = float(
            current_balance
        )  # Adjusted for new column

        # Start rows for transactions
        row = 4

        # Fetch Transactions
        transactions = ledger_filter(source=account, destination=account).order_by(
            "created_at"
        )

        # Add Transactions
        for i, tx in enumerate(transactions, start=1):
            if tx.transaction_type in [
                "INVOICE_PAYMENT",
                "ADD_ACC_BALANCE",
                "MISC_LOAN_CREATE",
            ]:
                credit = tx.amount
                debit = 0
                current_balance += credit
            else:
                credit = 0
                debit = tx.amount
                current_balance -= debit

            # Add transaction details to the sheet
            sheet.cell(row=row, column=start_col).value = i
            sheet.cell(row=row, column=start_col + 1).value = tx.created_at.strftime(
                "%Y-%m-%d %H:%M"
            )
            sheet.cell(row=row, column=start_col + 2).value = tx.reason
            sheet.cell(row=row, column=start_col + 3).value = (
                tx.get_transaction_type_display()
            )  # New column
            sheet.cell(row=row, column=start_col + 4).value = float(debit)
            sheet.cell(row=row, column=start_col + 5).value = float(credit)
            sheet.cell(row=row, column=start_col + 6).value = float(current_balance)
            row += 1

        # Add Total Row
        sheet.cell(row=row, column=start_col + 5).value = "Total"
        sheet.cell(row=row, column=start_col + 6).value = float(
            current_balance
        )  # Adjusted for new column

        # Move to the next account's column group (next 8 columns for tighter spacing)
        start_col += 8  # Adjusted spacing for new column

    # Save the workbook to a response object
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="bank_statements.xlsx"'
    workbook.save(response)

    return response


def generate_expense_report(request, start, end):
    # Filter expenses within the date range
    expenses = Ledger.objects.filter(
        transaction_type="MISC_EXPENSE", created_at__date__range=(start, end)
    )

    # Create a new workbook and select the active sheet
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Expense Report"

    # Define styles
    title_font = Font(size=14, bold=True, color="FFFFFF")
    header_font = Font(size=12, bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    bold_font = Font(size=11, bold=True)
    centered_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    title_fill = PatternFill("solid", fgColor="4F81BD")  # Blue background
    header_fill = PatternFill("solid", fgColor="A9C6E8")  # Light blue background
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    bold_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Helper function to style a cell
    def apply_cell_style(cell, font=None, alignment=None, fill=None, border=None):
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
        if fill:
            cell.fill = fill
        if border:
            cell.border = border

    # Add "Expense Report" Title
    current_row = 1
    sheet.merge_cells(
        start_row=current_row, start_column=1, end_row=current_row, end_column=8
    )  # Adjust to 7 columns now
    sheet[f"A{current_row}"] = f"Expense Report ({start} to {end})"
    apply_cell_style(
        sheet[f"A{current_row}"],
        font=title_font,
        alignment=centered_alignment,
        fill=title_fill,
    )

    # Add table headers
    current_row += 1
    expense_headers = [
        "Description",
        "Amount",
        "Source",
        "Category",
        "Vendor",
    ]
    for col, header in enumerate(expense_headers, start=1):
        cell = sheet.cell(row=current_row, column=col)
        cell.value = header
        apply_cell_style(
            cell,
            font=header_font,
            alignment=centered_alignment,
            fill=header_fill,
            border=thin_border,
        )

    # Dictionary to hold category totals
    category_totals = {}

    # Populate data rows and calculate total
    current_row += 1
    total_amount = 0

    for expense in expenses:
        row_data = [
            expense.reason,
            expense.amount,
            expense.source.get_name(),
            expense.expense_category.name,
            expense.destination.name,
        ]
        sheet.append(row_data)

        # Sum the expenses for each category
        if expense.expense_category.name not in category_totals:
            category_totals[expense.expense_category.name] = 0
        category_totals[expense.expense_category.name] += expense.amount

        # Apply styles to the row
        for col in range(1, 6):
            cell = sheet.cell(row=current_row, column=col)
            apply_cell_style(
                cell,
                font=normal_font,
                alignment=centered_alignment,
                border=thin_border,
            )

        # Add to the overall total amount
        total_amount += expense.amount

        current_row += 1  # Move to the next row

    # Add total row for all expenses
    sheet.append(["Total Expenses", total_amount, "", "", ""])
    apply_cell_style(
        sheet[f"B{current_row}"],
        font=bold_font,
        alignment=right_alignment,
        border=bold_border,
    )

    # Now add the category totals in a separate column (side by side)
    # Set starting row for Category Totals
    category_totals_row = 2  # Adjust this value to control placement

    # Add Category Totals Header (in column 7 and 8)
    sheet[f"G{category_totals_row}"] = "Category"
    sheet[f"H{category_totals_row}"] = "Total Amount"
    apply_cell_style(
        sheet[f"G{category_totals_row}"],
        font=bold_font,
        alignment=centered_alignment,
        fill=header_fill,
        border=bold_border,
    )
    apply_cell_style(
        sheet[f"H{category_totals_row}"],
        font=bold_font,
        alignment=centered_alignment,
        fill=header_fill,
        border=bold_border,
    )

    category_totals_row += 1

    # Add the category totals data on the right side (in columns G and H)
    for category, total in category_totals.items():
        sheet[f"G{category_totals_row}"] = category
        sheet[f"H{category_totals_row}"] = total

        apply_cell_style(
            sheet[f"G{category_totals_row}"],
            font=normal_font,
            alignment=left_alignment,
            border=thin_border,
        )
        apply_cell_style(
            sheet[f"H{category_totals_row}"],
            font=normal_font,
            alignment=right_alignment,
            border=thin_border,
        )

        category_totals_row += 1

    # Add Grand Total to the right of category totals (in column H)
    sheet.append(["", ""])
    apply_cell_style(
        sheet[f"G{category_totals_row}"],
        font=bold_font,
        alignment=centered_alignment,
        border=thin_border,
    )
    apply_cell_style(
        sheet[f"H{category_totals_row}"],
        font=bold_font,
        alignment=right_alignment,
        fill=title_fill,
        border=bold_border,
    )

    grand_total = sum(category_totals.values())
    sheet[f"H{category_totals_row}"].value = grand_total

    # Move to the next row
    category_totals_row += 1

    # Adjust column widths for both sides
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value and not isinstance(cell, MergedCell):  # Skip merged cells
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max(max_length + 2, 15)

    # Save the workbook to a response object
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="expense_report.xlsx"'
    wb.save(response)

    return response

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from django.http import HttpResponse

def yearly_report(request):
    # Get the selected year from the POST request
    year = request.POST.get('year') if request.method == 'POST' else None
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Projects"

    # Styles
    header_font = Font(size=12, bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    centered_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # Blue background for headers
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Add a title row with year information (row 1)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    sheet["A1"] = f"Yearly Report for {year if year else 'All Years'}"
    sheet["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    sheet["A1"].alignment = centered_alignment
    sheet["A1"].fill = header_fill

    # Column Headers for Projects (row 2)
    headers = ["Sr. No.", "Name", "Operating Expense", "Invoiced", "Received"]
    for col_num in range(1, 6):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 20

    # Add project data headers (row 2)
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # Add header for Category Totals side by side (columns G and H)
    sheet["G2"] = "Category"
    sheet["H2"] = "Total Amount"
    apply_cell_style(sheet["G2"], font=header_font, alignment=centered_alignment, fill=header_fill, border=thin_border)
    apply_cell_style(sheet["H2"], font=header_font, alignment=centered_alignment, fill=header_fill, border=thin_border)

    # Get Project Data and add it to the sheet (starting from row 3)
    row = 3
    for project in Project.objects.all():
        operating_expense = Ledger.objects.filter(
            project=project,
            created_at__year=year,
            transaction_type='CREATE_EXPENSE'
        ).aggregate(Sum('amount'))['amount__sum'] or 0

        invoiced = project.invoices.filter(created_at__year=year).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        received = project.invoices.filter(created_at__year=year, status='PAID').aggregate(Sum('total_amount'))['total_amount__sum'] or 0

        # Add data to project columns (A - E)
        sheet.cell(row=row, column=1).value = row - 2  # Sr. No.
        sheet.cell(row=row, column=2).value = project.project_name
        sheet.cell(row=row, column=3).value = float(operating_expense)
        sheet.cell(row=row, column=4).value = float(invoiced)
        sheet.cell(row=row, column=5).value = float(received)

        # Apply styling to project columns
        for col_num in range(1, 6):
            cell = sheet.cell(row=row, column=col_num)
            cell.alignment = centered_alignment
            cell.border = thin_border

        row += 1

    # Now, add category totals starting from row 3, columns G and H
    category_totals_row = 3  # This ensures the category totals start below the headers

    # Get the sum of amounts for each category
    category_totals = {}
    ledger_entries = Ledger.objects.filter(transaction_type__in=["CREATE_EXPENSE", "MISC_EXPENSE"], created_at__year=year)

    for entry in ledger_entries:
        category = entry.expense_category
        if category:
            if category not in category_totals:
                category_totals[category] = 0
            category_totals[category] += entry.amount

    # Add category totals to the sheet (starting from row 3, columns G and H)
    for category, total in category_totals.items():
        sheet.cell(row=category_totals_row, column=7).value = category.name  # Category name in column G
        sheet.cell(row=category_totals_row, column=8).value = total  # Total amount in column H

        # Apply styling to category columns
        apply_cell_style(sheet[f"G{category_totals_row}"], font=normal_font, alignment=left_alignment, border=thin_border)
        apply_cell_style(sheet[f"H{category_totals_row}"], font=normal_font, alignment=centered_alignment, border=thin_border)

        category_totals_row += 1

    # Adjust column widths automatically based on content
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Get column letter from the first cell
        for cell in col:
            if cell.value and not isinstance(cell, MergedCell):  # Skip merged cells
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max(max_length + 2, 15)

    # Save the workbook to a response object
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="projects_section.xlsx"'
    workbook.save(response)

    return response

# Helper function for styling
def apply_cell_style(cell, font, alignment, border, fill=None,):
    cell.font = font
    cell.alignment = alignment
    if fill:
        cell.fill = fill
    cell.border = border
